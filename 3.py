#!/usr/bin/env python3
"""
Extract data from campus placement PDF emails and populate Excel sheet
"""

import pdfplumber
import re
import openpyxl
import os
from openpyxl import Workbook

def extract_pdf_text(pdf_path):
    """Extract all text from a PDF file"""
    try:
        pdf = pdfplumber.open(pdf_path)
        text = ''
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                text += t + '\n'
        pdf.close()
        return text
    except Exception as e:
        print(f"Error reading {pdf_path}: {e}")
        return ""

def is_hackathon_or_internship(text, filename):
    """Check if PDF is hackathon or internship related (should be ignored)"""
    text_lower = text.lower()
    filename_lower = filename.lower()
    keywords = ['hackathon', 'internship', 'flipkart careers', 'hcl tech', 'career opportunities',
                'photography session', 'placement orientation', 'tcs national', 'wells fargo',
                'jaro education', 'odoo', 'grid 8.0', 'amplified', 'ai challenge']
    for kw in keywords:
        if kw in text_lower or kw in filename_lower:
            return True
    return False

def extract_company_name(text, filename):
    """Extract company name from PDF text or filename"""
    # Try from filename first
    match = re.search(r'Campus by\s+([^\.]+?)(?:\.pdf|\d+\.pdf)', filename, re.IGNORECASE)
    if match:
        name = match.group(1).strip()
        name = re.sub(r'\d+\.pdf$', '', name).strip()
        return name

    # Try from text
    patterns = [
        r'Campus by\s+([^\n]+)',
        r'Company Name\s*[:\n]\s*([^\n]+)',
        r'Company\s+Name\s*[:\n]\s*([^\n]+)',
    ]
    for p in patterns:
        m = re.search(p, text, re.IGNORECASE)
        if m:
            return m.group(1).strip()

    return "NA"

def extract_company_type(text):
    """Extract company type"""
    patterns = [
        r'Type of Company\s*[:\n]\s*([^\n]+)',
        r'Company Type\s*[:\n]\s*([^\n]+)',
    ]
    for p in patterns:
        m = re.search(p, text, re.IGNORECASE)
        if m:
            val = m.group(1).strip()
            val = re.sub(r'\s+', ' ', val)
            return val
    return "NA"

def extract_campus_date(text):
    """Extract campus date & time"""
    patterns = [
        r'Campus Date[^\n]*\n\s*([^\n]+)',
        r'Campus Date & Time\s*[:\n]\s*([^\n]+)',
        r'Campus Date & Time[^\n]*\n\s*([^\n]+)',
    ]
    for p in patterns:
        m = re.search(p, text, re.IGNORECASE)
        if m:
            val = m.group(1).strip()
            if val and val.lower() != 'will be declared soon':
                return val
            return "Will Be Declared Soon"
    return "NA"

def extract_last_reg_date(text):
    """Extract last registration date"""
    patterns = [
        r'Last Registration Date\s*[:\n]\s*([^\n]+)',
        r'Last Registration Date[^\n]*\n\s*([^\n]+)',
    ]
    for p in patterns:
        m = re.search(p, text, re.IGNORECASE)
        if m:
            val = m.group(1).strip()
            return val
    return "NA"

def extract_work_location(text):
    """Extract work location"""
    patterns = [
        r'Work Location\s*[:\n]\s*([^\n]+)',
        r'Work Location[^\n]*\n\s*([^\n]+)',
    ]
    for p in patterns:
        m = re.search(p, text, re.IGNORECASE)
        if m:
            val = m.group(1).strip()
            if val:
                return val
    return "NA"

def extract_job_profiles_and_packages(text):
    """Extract job profiles and their packages"""
    profiles = []
    packages = []

    # Find job description section
    job_section_match = re.search(
        r'Job Description\s+Job Profile\s+Package\s*\(CTC\)\s*(.+?)(?:Work Location|Skills|Selection|Eligibility|Job Description|$)',
        text, re.IGNORECASE | re.DOTALL
    )

    if job_section_match:
        jobs_text = job_section_match.group(1)

        # Pattern 1: Numbered list with package on same line
        pattern1 = r'(\d+\.\s*[^\n]+?)\s+(\d+\.?\d*\s*-\s*\d+\.?\d*\s*Lpa|\d+\.?\d*\s*Lpa|As per company|Negotiable|\d+\s*LPA)'
        matches = re.findall(pattern1, jobs_text, re.IGNORECASE)
        for m in matches:
            profile = m[0].strip()
            package = m[1].strip()
            profile = re.sub(r'^\d+\.\s*', '', profile).strip()
            profiles.append(profile)
            packages.append(package)

        # Pattern 2: Numbered list, package may be separate
        if not profiles:
            pattern2 = r'(\d+\.\s*[^\n]+)'
            matches = re.findall(pattern2, jobs_text)
            for m in matches:
                profile = re.sub(r'^\d+\.\s*', '', m).strip()
                profiles.append(profile)
                packages.append("NA")

        # Pattern 3: Bullet points or dash
        if not profiles:
            pattern3 = r'[•\-\*]\s*([^\n]+)'
            matches = re.findall(pattern3, jobs_text)
            for m in matches:
                profiles.append(m.strip())
                packages.append("NA")

    # If still no profiles, look for any job profile mentions
    if not profiles:
        profile_keywords = [
            r'(Software Developer|Data Analyst|Data Scientist|Intern|Trainee|Engineer|Developer|Analyst|'
            r'Business Development|HR|Sales|Marketing|DevOps|Full Stack|Backend|Frontend|'
            r'Project Coordinator|Quality Assurance|Designer|Content Writer|Technical)[^\n]{0,50}'
        ]
        for kw in profile_keywords:
            matches = re.findall(kw, text, re.IGNORECASE)
            for m in matches[:3]:
                profiles.append(m.strip()[:100])
                packages.append("NA")

    return profiles, packages

def extract_eligibility(text):
    """Extract eligibility criteria: 10th%, 12th%, CGPA, Backlog"""
    result = {
        'tenth': 'NA',
        'twelfth': 'NA',
        'cgpa': 'NA',
        'backlog': 'NA'
    }

    # Find eligibility section
    elig_section = re.search(
        r'Eligibility Parameters\s+Education Qualification\s+Eligibility Criteria\s*(.+?)(?:To Participate|Registration|$)',
        text, re.IGNORECASE | re.DOTALL
    )

    if elig_section:
        elig_text = elig_section.group(1)

        # 10th percentage
        m = re.search(r'(\d+\.?\d*)\s*%\s*-\s*10\s*th', elig_text, re.IGNORECASE)
        if m:
            result['tenth'] = m.group(1) + '%'

        # 12th percentage
        m = re.search(r'(\d+\.?\d*)\s*%\s*-\s*12\s*th', elig_text, re.IGNORECASE)
        if m:
            result['twelfth'] = m.group(1) + '%'

        # CGPA/CPI
        m = re.search(r'(\d+\.?\d*)\s*CPI', elig_text, re.IGNORECASE)
        if not m:
            m = re.search(r'(\d+\.?\d*)\s*CGPA', elig_text, re.IGNORECASE)
        if m:
            result['cgpa'] = m.group(1)

        # Backlog
        m = re.search(r'(No Active Backlog|Active Backlog|No Backlog)', elig_text, re.IGNORECASE)
        if m:
            result['backlog'] = m.group(1)

    # Alternative patterns in general text
    if result['tenth'] == 'NA':
        m = re.search(r'(\d+\.?\d*)\s*%\s*[-–]\s*10\s*(?:th|std)', text, re.IGNORECASE)
        if m:
            result['tenth'] = m.group(1) + '%'

    if result['twelfth'] == 'NA':
        m = re.search(r'(\d+\.?\d*)\s*%\s*[-–]\s*12\s*(?:th|std)', text, re.IGNORECASE)
        if m:
            result['twelfth'] = m.group(1) + '%'

    if result['cgpa'] == 'NA':
        m = re.search(r'(\d+\.?\d*)\s*(?:CPI|CGPA)\s*[&]?\s*[Aa]bove', text, re.IGNORECASE)
        if m:
            result['cgpa'] = m.group(1)

    if result['backlog'] == 'NA':
        m = re.search(r'(No Active Backlog|No Backlog|Active Backlog)', text, re.IGNORECASE)
        if m:
            result['backlog'] = m.group(1)

    return result

def process_pdf(pdf_path, sr_no):
    """Process a single PDF and return extracted data rows"""
    text = extract_pdf_text(pdf_path)

    if not text or len(text) < 100:
        return []

    # Skip hackathons and internships
    if is_hackathon_or_internship(text, pdf_path):
        print(f"  Skipping (hackathon/internship): {pdf_path}")
        return []

    company_name = extract_company_name(text, pdf_path)
    company_type = extract_company_type(text)
    campus_date = extract_campus_date(text)
    last_reg_date = extract_last_reg_date(text)
    work_location = extract_work_location(text)
    profiles, packages = extract_job_profiles_and_packages(text)
    eligibility = extract_eligibility(text)

    # If multiple job profiles, create separate rows
    rows = []
    if profiles and len(profiles) > 0:
        for i, (profile, package) in enumerate(zip(profiles, packages)):
            row = {
                'sr_no': sr_no + i,
                'company_name': company_name,
                'company_type': company_type,
                'campus_date': campus_date,
                'last_reg_date': last_reg_date,
                'job_profile': profile[:100] if profile else 'NA',
                'package': package if package else 'NA',
                'work_location': work_location,
                'tenth_percent': eligibility['tenth'],
                'twelfth_percent': eligibility['twelfth'],
                'cgpa': eligibility['cgpa'],
                'backlog': eligibility['backlog']
            }
            rows.append(row)
    else:
        # Single row with NA for job profile
        row = {
            'sr_no': sr_no,
            'company_name': company_name,
            'company_type': company_type,
            'campus_date': campus_date,
            'last_reg_date': last_reg_date,
            'job_profile': 'NA',
            'package': 'NA',
            'work_location': work_location,
            'tenth_percent': eligibility['tenth'],
            'twelfth_percent': eligibility['twelfth'],
            'cgpa': eligibility['cgpa'],
            'backlog': eligibility['backlog']
        }
        rows.append(row)

    return rows

def main():
    # Get all PDF files
    all_pdfs = [f for f in os.listdir('.') if f.endswith('.pdf')]

    # Filter out non-campus PDFs
    exclude_keywords = ['hackathon', 'flipkart', 'hcl', 'internship', 'career', 'photography',
                        'placement orientation', 'tcs national', 'wells fargo', 'jaro',
                        'odoo', 'grid', 'amplified', 'ai challenge', 'ww.pdf', 'consent',
                        'indian navy']

    campus_pdfs = []
    for pdf in all_pdfs:
        pdf_lower = pdf.lower()
        if not any(kw in pdf_lower for kw in exclude_keywords):
            campus_pdfs.append(pdf)

    campus_pdfs.sort()

    print(f"Found {len(campus_pdfs)} campus placement PDFs to process")

    all_rows = []
    sr_no = 1

    for pdf in campus_pdfs:
        print(f"Processing: {pdf}")
        rows = process_pdf(pdf, sr_no)
        all_rows.extend(rows)
        sr_no += len(rows)
        if rows:
            for r in rows:
                print(f"  Row {r['sr_no']}: {r['company_name']} - {r['job_profile']} - {r['package']}")

    # Write to Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Placement Data"

    # Headers
    headers = [
        'Sr No', 'Comapny name', 'Type of Company', 'Campus Date & Time',
        'Last Registration Date', 'Job Profile', 'Package', 'Work Location',
        '10th percentage', '12th percentage', 'cgpa', 'backlog'
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Data rows
    for row_idx, row in enumerate(all_rows, 2):
        ws.cell(row=row_idx, column=1, value=row['sr_no'])
        ws.cell(row=row_idx, column=2, value=row['company_name'])
        ws.cell(row=row_idx, column=3, value=row['company_type'])
        ws.cell(row=row_idx, column=4, value=row['campus_date'])
        ws.cell(row=row_idx, column=5, value=row['last_reg_date'])
        ws.cell(row=row_idx, column=6, value=row['job_profile'])
        ws.cell(row=row_idx, column=7, value=row['package'])
        ws.cell(row=row_idx, column=8, value=row['work_location'])
        ws.cell(row=row_idx, column=9, value=row['tenth_percent'])
        ws.cell(row=row_idx, column=10, value=row['twelfth_percent'])
        ws.cell(row=row_idx, column=11, value=row['cgpa'])
        ws.cell(row=row_idx, column=12, value=row['backlog'])

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    output_file = 'Record_filled.xlsx'
    wb.save(output_file)
    print(f"\n✅ Successfully saved {len(all_rows)} rows to {output_file}")

if __name__ == '__main__':
    main()