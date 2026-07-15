#!/usr/bin/env python3
"""
Extract data from campus placement PDF emails and populate Excel sheet
Improved version with better parsing
"""

import pdfplumber
import re
import openpyxl
import os
from openpyxl import Workbook

def extract_pdf_text(pdf_path):
    """Extract all text from a PDF file"""
    try:
        pdf = pdfplumber.open(pdf_path)
        text = ''
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                text += t + '\n'
        pdf.close()
        return text
    except Exception as e:
        print(f"Error reading {pdf_path}: {e}")
        return ""

def should_skip_pdf(text, filename):
    """Check if PDF should be skipped (hackathon, internship, non-campus)"""
    text_lower = text.lower()
    filename_lower = filename.lower()

    skip_conditions = [
        'hackathon' in filename_lower,
        'flipkart careers with grid' in text_lower,
        'hcl tech _ amplified' in text_lower,
        'internship program' in filename_lower,
        'career opportunities in indian navy' in text_lower,
        'photography session for placement' in filename_lower,
        'placement orientation' in filename_lower,
        'tcs national qualifier test' in filename_lower,
        'wells fargo' in filename_lower and 'internship' in text_lower,
        'jaro education corporate growth' in text_lower,
        'invitation to participate in odoo' in filename_lower,
        'career consent process' in filename_lower,
    ]

    return any(skip_conditions)

def extract_company_name(text, filename):
    """Extract company name"""
    # From filename
    match = re.search(r'Campus by\s+([^\.]+?)(?:\.pdf|\d+\.pdf)', filename, re.IGNORECASE)
    if match:
        name = match.group(1).strip()
        name = re.sub(r'\d+\.pdf$', '', name).strip()
        return name

    # From text
    patterns = [
        r'Campus by\s+([^\n]+)',
        r'Company Name\s*[:\n]\s*([^\n]+)',
        r'Company\s+Name\s*[:\n]\s*([^\n]+)',
    ]
    for p in patterns:
        m = re.search(p, text, re.IGNORECASE)
        if m:
            return m.group(1).strip()

    return "NA"

def extract_company_type(text):
    """Extract company type"""
    patterns = [
        r'Type of Company\s*[:\n]\s*([^\n]+)',
        r'Company Type\s*[:\n]\s*([^\n]+)',
    ]
    for p in patterns:
        m = re.search(p, text, re.IGNORECASE)
        if m:
            val = m.group(1).strip()
            val = re.sub(r'\s+', ' ', val)
            return val
    return "NA"

def extract_campus_date(text):
    """Extract campus date & time"""
    # Find the campus date section
    patterns = [
        r'Campus Date & Time\s*[:\n]\s*([^\n]+)',
        r'Campus Date & Time[^\n]*\n\s*([^\n]+)',
    ]
    for p in patterns:
        m = re.search(p, text, re.IGNORECASE)
        if m:
            val = m.group(1).strip()
            if val and val.lower() != 'will be declared soon':
                return val
            return "Will Be Declared Soon"
    return "NA"

def extract_last_reg_date(text):
    """Extract last registration date"""
    patterns = [
        r'Last Registration Date\s*[:\n]\s*([^\n]+)',
        r'Last Registration Date[^\n]*\n\s*([^\n]+)',
    ]
    for p in patterns:
        m = re.search(p, text, re.IGNORECASE)
        if m:
            val = m.group(1).strip()
            return val
    return "NA"

def extract_work_location(text):
    """Extract work location"""
    patterns = [
        r'Work Location\s*[:\n]\s*([^\n]+)',
        r'Work Location[^\n]*\n\s*([^\n]+)',
    ]
    for p in patterns:
        m = re.search(p, text, re.IGNORECASE)
        if m:
            val = m.group(1).strip()
            if val:
                return val
    return "NA"

def extract_job_profiles_and_packages(text):
    """Extract job profiles and their packages - improved parsing"""
    profiles = []
    packages = []

    # Find job description section
    job_section_match = re.search(
        r'Job Description\s+Job Profile\s+Package\s*\(CTC\)\s*(.+?)(?:Work Location|Skills|Selection|Eligibility|Job Description|$)',
        text, re.IGNORECASE | re.DOTALL
    )

    if job_section_match:
        jobs_text = job_section_match.group(1)

        # Pattern: Numbered list with profile and package
        # Handle formats like:
        # "1. Dot Net Developer Dot Net Developer : 4.0 to 5.0"
        # "2. React Developer LPA"
        # "1. Software Developer 6.00 - 6.50 Lpa"
        # "1. Project Coordinator 1) Project Coordinator - 2.64 LPA"

        # Split by numbered items
        items = re.split(r'\n\s*\d+\.\s*', jobs_text)
        for item in items:
            item = item.strip()
            if not item or len(item) < 3:
                continue

            # Try to extract package from the item
            package = "NA"
            profile = item

            # Package patterns
            pkg_patterns = [
                r'(\d+\.?\d*\s*-\s*\d+\.?\d*\s*Lpa)',
                r'(\d+\.?\d*\s*Lpa)',
                r'(CTC\s*:\s*[\d\.]+\s*LPA)',
                r'(\d+\s*LPA)',
                r'(As per company|Negotiable)',
            ]
            for pp in pkg_patterns:
                pm = re.search(pp, item, re.IGNORECASE)
                if pm:
                    package = pm.group(1).strip()
                    # Remove package from profile
                    profile = re.sub(pp, '', item, flags=re.IGNORECASE).strip()
                    break

            # Clean profile
            profile = re.sub(r'^\d+\)\s*', '', profile).strip()
            profile = re.sub(r'\s+', ' ', profile).strip()

            if profile and len(profile) > 2:
                profiles.append(profile[:100])
                packages.append(package)

        # If still no profiles, try alternative parsing
        if not profiles:
            # Look for lines with job titles
            lines = jobs_text.split('\n')
            for line in lines:
                line = line.strip()
                if not line or len(line) < 3:
                    continue
                # Skip lines that are just numbers or bullets
                if re.match(r'^[\d\.\-\•\s]+$', line):
                    continue
                profiles.append(line[:100])
                packages.append("NA")

    # If still no profiles, try to find job profiles elsewhere
    if not profiles:
        # Look for numbered job profiles in eligibility or other sections
        pattern = r'(\d+\.\s*(?:Software|Data|Developer|Engineer|Analyst|Intern|Trainee|HR|Sales|Marketing|DevOps|Backend|Frontend|Full Stack|Project|Quality|Designer|Content|Technical|Business|UI/UX|AI/ML|Cloud|System|Network|Database|Security|Mobile|Web|Dot Net|React|Python|Java|Node)[^\n]{5,80})'
        matches = re.findall(pattern, text, re.IGNORECASE)
        for m in matches[:5]:
            profile = re.sub(r'^\d+\.\s*', '', m).strip()
            profiles.append(profile[:100])
            packages.append("NA")

    return profiles, packages

def extract_eligibility(text):
    """Extract eligibility criteria"""
    result = {
        'tenth': 'NA',
        'twelfth': 'NA',
        'cgpa': 'NA',
        'backlog': 'NA'
    }

    # Find eligibility section
    elig_section = re.search(
        r'Eligibility Parameters\s+Education Qualification\s+Eligibility Criteria\s*(.+?)(?:To Participate|Registration|$)',
        text, re.IGNORECASE | re.DOTALL
    )

    if elig_section:
        elig_text = elig_section.group(1)

        m = re.search(r'(\d+\.?\d*)\s*%\s*-\s*10\s*th', elig_text, re.IGNORECASE)
        if m:
            result['tenth'] = m.group(1) + '%'

        m = re.search(r'(\d+\.?\d*)\s*%\s*-\s*12\s*th', elig_text, re.IGNORECASE)
        if m:
            result['twelfth'] = m.group(1) + '%'

        m = re.search(r'(\d+\.?\d*)\s*CPI', elig_text, re.IGNORECASE)
        if not m:
            m = re.search(r'(\d+\.?\d*)\s*CGPA', elig_text, re.IGNORECASE)
        if m:
            result['cgpa'] = m.group(1)

        m = re.search(r'(No Active Backlog|Active Backlog|No Backlog)', elig_text, re.IGNORECASE)
        if m:
            result['backlog'] = m.group(1)

    # Alternative patterns in general text
    if result['tenth'] == 'NA':
        m = re.search(r'(\d+\.?\d*)\s*%\s*[-–]\s*10\s*(?:th|std)', text, re.IGNORECASE)
        if m:
            result['tenth'] = m.group(1) + '%'

    if result['twelfth'] == 'NA':
        m = re.search(r'(\d+\.?\d*)\s*%\s*[-–]\s*12\s*(?:th|std)', text, re.IGNORECASE)
        if m:
            result['twelfth'] = m.group(1) + '%'

    if result['cgpa'] == 'NA':
        m = re.search(r'(\d+\.?\d*)\s*(?:CPI|CGPA)\s*[&]?\s*[Aa]bove', text, re.IGNORECASE)
        if m:
            result['cgpa'] = m.group(1)

    if result['backlog'] == 'NA':
        m = re.search(r'(No Active Backlog|No Backlog|Active Backlog)', text, re.IGNORECASE)
        if m:
            result['backlog'] = m.group(1)

    return result

def process_pdf(pdf_path, sr_no):
    """Process a single PDF"""
    text = extract_pdf_text(pdf_path)

    if not text or len(text) < 100:
        return []

    if should_skip_pdf(text, pdf_path):
        print(f"  Skipping (hackathon/internship/non-campus): {pdf_path}")
        return []

    company_name = extract_company_name(text, pdf_path)
    company_type = extract_company_type(text)
    campus_date = extract_campus_date(text)
    last_reg_date = extract_last_reg_date(text)
    work_location = extract_work_location(text)
    profiles, packages = extract_job_profiles_and_packages(text)
    eligibility = extract_eligibility(text)

    rows = []
    if profiles and len(profiles) > 0:
        for i, (profile, package) in enumerate(zip(profiles, packages)):
            row = {
                'sr_no': sr_no + i,
                'company_name': company_name,
                'company_type': company_type,
                'campus_date': campus_date,
                'last_reg_date': last_reg_date,
                'job_profile': profile[:100] if profile else 'NA',
                'package': package if package else 'NA',
                'work_location': work_location,
                'tenth_percent': eligibility['tenth'],
                'twelfth_percent': eligibility['twelfth'],
                'cgpa': eligibility['cgpa'],
                'backlog': eligibility['backlog']
            }
            rows.append(row)
    else:
        row = {
            'sr_no': sr_no,
            'company_name': company_name,
            'company_type': company_type,
            'campus_date': campus_date,
            'last_reg_date': last_reg_date,
            'job_profile': 'NA',
            'package': 'NA',
            'work_location': work_location,
            'tenth_percent': eligibility['tenth'],
            'twelfth_percent': eligibility['twelfth'],
            'cgpa': eligibility['cgpa'],
            'backlog': eligibility['backlog']
        }
        rows.append(row)

    return rows

def main():
    all_pdfs = [f for f in os.listdir('.') if f.endswith('.pdf')]
    all_pdfs.sort()

    print(f"Found {len(all_pdfs)} total PDF files")

    all_rows = []
    sr_no = 1

    for pdf in all_pdfs:
        print(f"Processing: {pdf}")
        rows = process_pdf(pdf, sr_no)
        if rows:
            all_rows.extend(rows)
            sr_no += len(rows)
            for r in rows:
                print(f"  Row {r['sr_no']}: {r['company_name']} - {r['job_profile']} - {r['package']}")

    # Write to Excel (overwrite original)
    wb = Workbook()
    ws = wb.active
    ws.title = "Placement Data"

    headers = [
        'Sr No', 'Comapny name', 'Type of Company', 'Campus Date & Time',
        'Last Registration Date', 'Job Profile', 'Package', 'Work Location',
        '10th percentage', '12th percentage', 'cgpa', 'backlog'
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    for row_idx, row in enumerate(all_rows, 2):
        ws.cell(row=row_idx, column=1, value=row['sr_no'])
        ws.cell(row=row_idx, column=2, value=row['company_name'])
        ws.cell(row=row_idx, column=3, value=row['company_type'])
        ws.cell(row=row_idx, column=4, value=row['campus_date'])
        ws.cell(row=row_idx, column=5, value=row['last_reg_date'])
        ws.cell(row=row_idx, column=6, value=row['job_profile'])
        ws.cell(row=row_idx, column=7, value=row['package'])
        ws.cell(row=row_idx, column=8, value=row['work_location'])
        ws.cell(row=row_idx, column=9, value=row['tenth_percent'])
        ws.cell(row=row_idx, column=10, value=row['twelfth_percent'])
        ws.cell(row=row_idx, column=11, value=row['cgpa'])
        ws.cell(row=row_idx, column=12, value=row['backlog'])

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column].width = adjusted_width

    output_file = 'Record.xlsx'
    wb.save(output_file)
    print(f"\n✅ Successfully saved {len(all_rows)} rows to {output_file}")

if __name__ == '__main__':
    main()